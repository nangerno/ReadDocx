from docx import Document
import openpyxl

# Load the Word document
doc = Document('file.docx')

current_course = None
current_lesson = None
current_quizz = None
current_aufgabe = None
current_frage = None
current_antwort = None
current_erklarung = None
current_options = []

# Create a new Excel workbook and worksheet
workbook = openpyxl.Workbook()
worksheet = workbook.active

# Write headers to the worksheet
font = openpyxl.styles.Font(color='FF0000', bold=True)
worksheet['A1'] = 'Course'
worksheet['A1'].font = font
worksheet['B1'] = 'Lesson'
worksheet['B1'].font = font
worksheet['C1'] = 'Question'
worksheet['C1'].font = font
worksheet['D1'] = 'Task'
worksheet['D1'].font = font
worksheet['E1'] = 'Ask'
worksheet['E1'].font = font
worksheet['F1'] = 'Problem'
worksheet['F1'].font = font
worksheet['G1'] = 'Answer'
worksheet['G1'].font = font
# Start writing data from the second row
row = 2 

for paragraph in doc.paragraphs:
    text = paragraph.text.strip()

    if text.startswith("Course:"):
        # New course found
        current_course = text[7:].strip()
        current_lesson = None
        current_quizz = None
        current_aufgabe = None
        current_frage = None
        current_antwort = None
        current_erklarung = None
        current_options = []
        worksheet.cell(row=row, column=1, value=current_course)
            
    elif text.startswith("Lesson:"):
        # New lesson found
        current_lesson = text[7:].strip()
        current_quizz = None
        current_aufgabe = None
        current_frage = None
        current_antwort = None
        current_erklarung = None
        current_options = []
        worksheet.cell(row=row, column=2, value=current_lesson)

    elif text.startswith("Quizz"):
        # New quizz found
        current_quizz = text
        current_aufgabe = None
        current_frage = None
        current_antwort = None
        current_erklarung = None
        current_options = []
        print(current_quizz)
        worksheet.cell(row=row, column=3, value=current_quizz)

    elif text.startswith("Aufgabe"):
        # New question found
        if current_aufgabe is not None:
            # Print the previous question's details

            # print(current_aufgabe)
            # print(f"Frage: {current_frage}")
            # print(f"{', '.join(current_options)}")
            # print(f"Antwort: {current_antwort}")
            # print()

            worksheet.cell(row=row, column=4, value=current_aufgabe)
            worksheet.cell(row=row, column=5, value=current_frage)
            worksheet.cell(row=row, column=6, value=', '.join(current_options))
            worksheet.cell(row=row, column=7, value=current_antwort)
            worksheet.cell(row=row, column=8, value=current_erklarung)
            row += 1

        # Extract the new question number
        current_aufgabe = text
        current_frage = None
        current_antwort = None
        current_erklarung = None
        current_options = []

    elif text.startswith("Frage:"):
        current_frage = text[6:].strip()

    elif text.startswith("Antwort:"):
        current_antwort = text[8:].strip()
    
    elif text.startswith("Erklärung:"):
        current_erklarung = text[10:].strip()

    else:
        # Check for options (A, B, C, D
        if text.startswith(('A)', 'B)', 'C)', 'D)', 'E)')):
            current_options.append(text.strip())
# Print the last question's details
if current_aufgabe is not None:
    worksheet.cell(row=row, column=3, value=current_quizz)
    worksheet.cell(row=row, column=4, value=current_aufgabe)
    worksheet.cell(row=row, column=5, value=current_frage)
    worksheet.cell(row=row, column=6, value=', '.join(current_options))
    worksheet.cell(row=row, column=7, value=current_antwort)
    worksheet.cell(row=row, column=8, value=current_erklarung)
print("------successfully------") 
workbook.save('result.xlsx')
